from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages 
from django.db import transaction 
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.db.models import Q, Sum, Avg, Count
from django.db.models.functions import TruncMonth
from django.utils import timezone
from django.contrib.auth.models import User
import json, csv, openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from decimal import Decimal
from functools import wraps

# Imports dos nossos modelos e forms refatorados
from .models import (
    LotePagamentoComissao, MetaVenda
)
from .forms import (
    TransacaoPagamentoForm, AnexoLoteForm, MetaVendaForm
)
# Imports das apps 'vendas' e 'common' (das quais dependemos)
from vendas.models import Venda
from common.models import Produto

# ---
# VISTO QUE ESTA APP DEPENDE DAS VIEWS DE 'VENDAS', IMPORTAMOS AS SUAS FUNÇÕES HELPER
# (Isto é uma dependência, mas é aceitável na arquitetura do Django)
# ---
from vendas.views import _get_user_permissions, _get_vendas_filtradas, gestor_ou_admin_required

# ---
# FUNÇÃO HELPER DESTA APP
# ---

def _get_lotes_filtrados(request):
    """
    Função auxiliar que lê os filtros do request (GET) e retorna
    o queryset de Lotes já filtrado e com as permissões corretas.
    """
    perms = _get_user_permissions(request.user)
    
    # 1. Queryset Base
    if perms['is_vendedor'] and not (perms['is_admin'] or perms['is_gestor'] or perms['is_financeiro']):
        lotes_qs = LotePagamentoComissao.objects.filter(vendedor=request.user)
    else:
        lotes_qs = LotePagamentoComissao.objects.all()

    # 2. Captura dos Parâmetros de Filtro
    query_vendedor = request.GET.get('vendedor', '')
    query_status = request.GET.get('status', '')
    query_data_inicio = request.GET.get('data_inicio', '')
    query_data_fim = request.GET.get('data_fim', '')

    # 3. Aplicação dos Filtros
    if (perms['is_admin'] or perms['is_gestor'] or perms['is_financeiro']) and query_vendedor:
        lotes_qs = lotes_qs.filter(vendedor_id=query_vendedor)
    if query_status:
        lotes_qs = lotes_qs.filter(status=query_status)
    if query_data_inicio:
        lotes_qs = lotes_qs.filter(data_fechamento__gte=query_data_inicio)
    if query_data_fim:
        try:
            data_fim_dt = datetime.strptime(query_data_fim, '%Y-%m-%d').date() + timedelta(days=1)
            lotes_qs = lotes_qs.filter(data_fechamento__lt=data_fim_dt)
        except (ValueError, TypeError):
            pass
            
    # 4. Otimiza a query e retorna
    return lotes_qs.select_related('vendedor', 'responsavel_fechamento').order_by('-data_fechamento')

# ---
# VIEWS DO MÓDULO DE COMISSÕES (Movidas de vendas/views.py)
# ---

@login_required
def comissoes_dashboard_graficos(request):
    """ Dashboard Gráfico de Comissões """
    perms = _get_user_permissions(request.user)
    if not (perms['is_admin'] or perms['is_gestor'] or perms['is_financeiro'] or perms['is_vendedor']):
        messages.error(request, "Você não tem permissão para aceder a esta página.")
        return redirect('dashboard')

    # 1. Reusa a lógica de filtros de Vendas (importada de vendas.views)
    vendas_filtradas = _get_vendas_filtradas(request)

    # 2. Filtro ADICIONAL
    comissoes_filtradas = vendas_filtradas.filter(status_pagamento='aprovado')

    # 3. Cálculo de KPIs de Comissão
    kpis = comissoes_filtradas.aggregate(
        total_comissoes=Sum('comissao_calculada_final'),
        ticket_medio_comissao=Avg('comissao_calculada_final'),
        contagem_vendas_comissionadas=Count('id')
    )

    # 4. Dados para Gráficos
    comissoes_por_vendedor = comissoes_filtradas.values('vendedor__username').annotate(total=Sum('comissao_calculada_final')).order_by('-total')
    comissoes_por_vendedor_payload = [{'vendedor__username': c['vendedor__username'], 'total': float(c['total'] or 0)} for c in comissoes_por_vendedor]
    comissoes_por_produto = comissoes_filtradas.values('produto__nome').annotate(total=Sum('comissao_calculada_final')).order_by('-total')
    comissoes_por_produto_payload = [{'produto__nome': p['produto__nome'], 'total': float(p['total'] or 0)} for p in comissoes_por_produto]
    comissoes_por_mes = comissoes_filtradas.annotate(mes=TruncMonth('data_venda')).values('mes').annotate(total=Sum('comissao_calculada_final')).order_by('mes')
    labels_linha = [c['mes'].strftime('%b/%Y') for c in comissoes_por_mes]
    data_linha = [float(c['total'] or 0) for c in comissoes_por_mes]

    # 5. Contexto para Filtros
    todos_vendedores = User.objects.filter(is_superuser=False, is_active=True).order_by('username')
    todos_produtos = Produto.objects.all().order_by('nome')
    today_dt = datetime.now().date()
    default_start_str = (today_dt - timedelta(days=30)).strftime('%Y-%m-%d')
    default_end_str = today_dt.strftime('%Y-%m-%d')

    context = {
        'perms': perms,
        'kpi_total_comissoes': kpis.get('total_comissoes') or 0,
        'kpi_media_comissao': kpis.get('ticket_medio_comissao') or 0,
        'kpi_contagem_comissoes': kpis.get('contagem_vendas_comissionadas') or 0,
        'comissoes_por_vendedor_json': json.dumps(comissoes_por_vendedor_payload),
        'comissoes_por_produto_json': json.dumps(comissoes_por_produto_payload),
        'comissoes_por_mes_labels_json': json.dumps(labels_linha),
        'comissoes_por_mes_data_json': json.dumps(data_linha),
        'todos_vendedores': todos_vendedores,
        'todos_produtos': todos_produtos,
        'query_data_inicio': request.GET.get('data_inicio', default_start_str),
        'query_data_fim': request.GET.get('data_fim', default_end_str),
    }
    return render(request, 'comissoes/dashboard_graficos.html', context)


@login_required
def comissoes_historico_lotes(request):
    """ Histórico de Lotes com Filtros """
    perms = _get_user_permissions(request.user)
    if not (perms['is_admin'] or perms['is_gestor'] or perms['is_financeiro'] or perms['is_vendedor']):
         messages.error(request, "Você não tem permissão para aceder a esta página.")
         return redirect('dashboard')

    # 1. USA A FUNÇÃO HELPER LOCAL
    lotes_pagos = _get_lotes_filtrados(request)

    # 2. Contexto para os Dropdowns
    vendedores_com_lotes = User.objects.filter(
        id__in=LotePagamentoComissao.objects.values_list('vendedor_id', flat=True).distinct()
    ).order_by('username')

    context = {
        'lotes_pagos': lotes_pagos,
        'perms': perms,
        'todos_vendedores': vendedores_com_lotes,
        'status_lote_choices': LotePagamentoComissao.STATUS_CHOICES,
    }
    return render(request, 'comissoes/historico_lotes.html', context)


@login_required
@transaction.atomic
def comissoes_fechamento(request):
    """ Página 'Pagar Comissões' (Fecho Manual por Período/Vendedor) """
    perms = _get_user_permissions(request.user)
    if not (perms['is_admin'] or perms['is_gestor'] or perms['is_financeiro']):
        messages.error(request, "Você não tem permissão para aceder a esta página.")
        return redirect('dashboard')
    
    vendas_para_pagar = None
    resumo_por_vendedor = None
    total_geral_comissao = 0
    
    vendedores_com_pendencias = User.objects.filter(
        is_superuser=False, is_active=True, 
        vendas__status_pagamento='aprovado', 
        vendas__lote_pagamento__isnull=True
    ).distinct().order_by('username')

    if request.method == 'GET' and 'data_inicio' in request.GET and 'data_fim' in request.GET:
        # ... (Lógica GET permanece a mesma) ...
        data_inicio = request.GET.get('data_inicio')
        data_fim = request.GET.get('data_fim')
        query_vendedor = request.GET.get('vendedor', '')
        
        if data_inicio and data_fim: 
            try:
                data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d').date() + timedelta(days=1)
            except (ValueError, TypeError):
                data_fim_dt = datetime.now().date() + timedelta(days=1)

            vendas_para_pagar = Venda.objects.filter(
                status_pagamento='aprovado',
                lote_pagamento__isnull=True,
                data_venda__gte=data_inicio,
                data_venda__lt=data_fim_dt 
            )
            
            if query_vendedor:
                vendas_para_pagar = vendas_para_pagar.filter(vendedor_id=query_vendedor)
            
            resumo_por_vendedor = vendas_para_pagar.values('vendedor__username') \
                                              .annotate(total_comissao=Sum('comissao_calculada_final'),
                                                        contagem=Count('id')) \
                                              .order_by('vendedor__username')
            total_geral_comissao = resumo_por_vendedor.aggregate(total_geral=Sum('total_comissao'))['total_geral'] or 0

    elif request.method == 'POST':
        # ... (Lógica POST permanece a mesma, usando o modelo LotePagamentoComissao local) ...
        data_inicio = request.POST.get('data_inicio')
        data_fim = request.POST.get('data_fim')
        query_vendedor = request.POST.get('vendedor', '')

        try:
            data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d').date() + timedelta(days=1)
        except (ValueError, TypeError):
            data_fim_dt = datetime.now().date() + timedelta(days=1)

        vendas_para_pagar = Venda.objects.filter(
            status_pagamento='aprovado',
            lote_pagamento__isnull=True,
            data_venda__gte=data_inicio,
            data_venda__lt=data_fim_dt
        )
        
        vendedores_a_processar_ids = []
        if query_vendedor:
            vendedores_a_processar_ids = [query_vendedor]
            vendas_para_pagar = vendas_para_pagar.filter(vendedor_id=query_vendedor)
        else:
            vendedores_a_processar_ids = vendas_para_pagar.values_list('vendedor_id', flat=True).distinct()

        if not vendas_para_pagar.exists():
            messages.error(request, "Nenhuma venda encontrada para fechar no período selecionado.")
            return redirect('comissoes_fechamento')

        lotes_criados_count = 0
        for vendor_id in vendedores_a_processar_ids:
            vendas_do_vendedor = vendas_para_pagar.filter(vendedor_id=vendor_id)
            total_do_vendedor = vendas_do_vendedor.aggregate(total=Sum('comissao_calculada_final'))['total'] or 0
            
            if total_do_vendedor > 0:
                novo_lote = LotePagamentoComissao.objects.create(
                    vendedor_id=vendor_id, 
                    periodo_inicio=data_inicio,
                    periodo_fim=data_fim,
                    responsavel_fechamento=request.user,
                    status='pendente',
                    total_comissoes=total_do_vendedor
                )
                vendas_do_vendedor.update(lote_pagamento=novo_lote)
                lotes_criados_count += 1
        
        if lotes_criados_count > 0:
            messages.success(request, f"{lotes_criados_count} lote(s) de pagamento criado(s) com sucesso.")
            return redirect('comissoes_historico')
        else:
            messages.error(request, "Nenhuma comissão a pagar encontrada para os filtros selecionados.")
            return redirect('comissoes_fechamento')

    context = {
        'vendas_para_pagar': vendas_para_pagar,
        'resumo_por_vendedor': resumo_por_vendedor,
        'total_geral_comissao': total_geral_comissao,
        'todos_vendedores': vendedores_com_pendencias,
    }
    return render(request, 'comissoes/comissoes_fechamento.html', context)

@login_required
@transaction.atomic
def comissoes_lote_detalhe(request, lote_id):
    """ Página de Detalhe do Lote """
    perms = _get_user_permissions(request.user)
    lote = get_object_or_404(LotePagamentoComissao, id=lote_id)

    if perms['is_vendedor'] and not (perms['is_admin'] or perms['is_gestor'] or perms['is_financeiro']):
        if lote.vendedor != request.user:
             messages.error(request, "Você não tem permissão para visualizar este lote.")
             return redirect('comissoes_historico')
             
    transacao_form = TransacaoPagamentoForm(lote=lote)
    anexo_lote_form = AnexoLoteForm() 

    if request.method == 'POST':
        if perms['is_vendedor'] and not (perms['is_admin'] or perms['is_gestor'] or perms['is_financeiro']):
            messages.error(request, "Você não tem permissão para esta ação.")
            return redirect('comissoes_lote_detalhe', lote_id=lote.id)

        acao = request.POST.get('acao')

        if acao == 'add_transacao':
            transacao_form = TransacaoPagamentoForm(request.POST, request.FILES, lote=lote)
            if transacao_form.is_valid():
                transacao = transacao_form.save(commit=False)
                transacao.lote = lote
                transacao.responsavel_pagamento = request.user
                if 'anexo_comprovativo_pagamento' in request.FILES:
                    transacao.descricao = request.FILES['anexo_comprovativo_pagamento'].name
                transacao.save()
                messages.success(request, "Pagamento registado com sucesso.")
                return redirect('comissoes_lote_detalhe', lote_id=lote.id)
        
        elif acao == 'add_anexo_nf':
            anexo_lote_form = AnexoLoteForm(request.POST, request.FILES)
            if anexo_lote_form.is_valid():
                anexo = anexo_lote_form.save(commit=False)
                anexo.lote = lote
                if 'anexo_nf_vendedor' in request.FILES:
                    anexo.descricao = request.FILES['anexo_nf_vendedor'].name
                anexo.save()
                messages.success(request, "Nota Fiscal do Vendedor anexada com sucesso.")
                return redirect('comissoes_lote_detalhe', lote_id=lote.id)

    vendas_no_lote = lote.vendas.all().order_by('data_venda')
    transacoes_do_lote = lote.transacoes_pagamento.all().order_by('data_pagamento')
    anexos_do_lote = lote.anexos_lote.all().order_by('data_upload')

    context = {
        'lote': lote,
        'vendas_no_lote': vendas_no_lote,
        'transacoes_do_lote': transacoes_do_lote,
        'anexos_do_lote': anexos_do_lote, 
        'transacao_form': transacao_form, 
        'anexo_lote_form': anexo_lote_form,
        'perms': perms,
    }
    return render(request, 'comissoes/comissoes_lote_detalhe.html', context)


# ---
# VIEWS DE EXPORTAÇÃO (Movidas de vendas/views.py)
# ---

@login_required
def export_lotes_csv(request):
    lotes = _get_lotes_filtrados(request)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="relatorio_lotes_comissao.csv"'
    writer = csv.writer(response, delimiter=';')
    
    headers = ['Lote ID', 'Data Fechamento', 'Vendedor', 'Responsável (Fechou)', 
               'Período Início', 'Período Fim', 
               'Total Devido', 'Total Pago', 'Status']
    writer.writerow(headers)
    
    for lote in lotes:
        writer.writerow([
            lote.id, 
            lote.data_fechamento.strftime('%Y-%m-%d %H:%M'), 
            lote.vendedor.username, 
            lote.responsavel_fechamento.username,
            lote.periodo_inicio,
            lote.periodo_fim,
            lote.total_comissoes,
            lote.total_pago_efetivamente,
            lote.get_status_display()
        ])
    return response

@login_required
def export_lotes_xlsx(request):
    lotes = _get_lotes_filtrados(request)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lotes de Comissão"
    
    headers = ['Lote ID', 'Data Fechamento', 'Vendedor', 'Responsável (Fechou)', 
               'Período Início', 'Período Fim', 
               'Total Devido (R$)', 'Total Pago (R$)', 'Status']
    ws.append(headers)
    
    for lote in lotes:
        ws.append([
            lote.id, 
            lote.data_fechamento.replace(tzinfo=None), 
            lote.vendedor.username, 
            lote.responsavel_fechamento.username,
            lote.periodo_inicio,
            lote.periodo_fim,
            float(lote.total_comissoes or 0),
            float(lote.total_pago_efetivamente or 0),
            lote.get_status_display()
        ])
        
    for col_num, header in enumerate(headers, 1): 
        ws.column_dimensions[get_column_letter(col_num)].width = 22
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="relatorio_lotes_comissao.xlsx"'
    wb.save(response)
    return response

# ---
# VIEWS DE GESTÃO DE METAS (Movidas de vendas/views.py)
# ---

@login_required
@gestor_ou_admin_required
def lista_metas(request):
    """ (R)ead: Lista todas as metas (passadas, presentes, futuras) """
    metas = MetaVenda.objects.all().select_related('vendedor', 'grupo').order_by('-data_inicio')
    context = {
        'metas': metas
    }
    return render(request, 'metas/lista_metas.html', context)

@login_required
@gestor_ou_admin_required
def criar_meta(request):
    """ (C)reate: Formulário para criar nova meta """
    if request.method == 'POST':
        form = MetaVendaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Nova meta criada com sucesso.")
            return redirect('lista_metas')
    else:
        form = MetaVendaForm()
    
    context = {
        'form': form,
        'form_title': 'Criar Nova Meta'
    }
    return render(request, 'metas/meta_form.html', context)

@login_required
@gestor_ou_admin_required
def editar_meta(request, meta_id):
    """ (U)pdate: Formulário para editar meta existente """
    meta = get_object_or_404(MetaVenda, id=meta_id)
    if request.method == 'POST':
        form = MetaVendaForm(request.POST, instance=meta)
        if form.is_valid():
            form.save()
            messages.success(request, "Meta atualizada com sucesso.")
            return redirect('lista_metas')
    else:
        form = MetaVendaForm(instance=meta)
    
    context = {
        'form': form,
        'form_title': f"Editar Meta ({meta})"
    }
    return render(request, 'metas/meta_form.html', context)

@login_required
@gestor_ou_admin_required
def apagar_meta(request, meta_id):
    """ (D)elete: Confirmação e exclusão de meta """
    meta = get_object_or_404(MetaVenda, id=meta_id)
    if request.method == 'POST':
        meta.delete()
        messages.success(request, "Meta apagada com sucesso.")
        return redirect('lista_metas')
    
    context = {
        'meta': meta
    }
    return render(request, 'metas/meta_confirm_delete.html', context)