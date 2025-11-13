# Em: vendas/views.py (Atualizado)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages 
from django.db import transaction 
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.db.models import Q, Sum, Avg, Count
from django.db.models.functions import TruncMonth
from django.utils import timezone
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
import json, csv, openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from decimal import Decimal
from functools import wraps

# --- (ATUALIZADO) Imports dos Modelos ---
from .models import Venda, AnexoVenda
from common.models import Cliente, Produto
from comissoes.models import (
    RegraComissaoVendedor, MetaVenda
)

# --- (ATUALIZADO) Imports dos Forms ---
from .forms import (
    VendaForm, AnexoForm,
    VendaStatusAdminForm, VendaStatusVendedorForm, 
    VendaStatusFinanceiroForm, VendaStatusAdvogadoForm, 
    VendaEditForm
)
from common.forms import ClienteForm, ClienteEditForm
# (Forms de comissões e metas foram removidos)


# ---
# SEÇÃO 1: FUNÇÕES AUXILIARES (LÓGICA INTERNA)
# ---

# (Estas funções helper permanecem aqui, pois são a "base" da lógica de negócio)

def _get_user_permissions(user):
    """ Verifica os grupos do utilizador e retorna um dicionário de flags. """
    return {
        'is_admin': user.is_superuser,
        'is_gestor': user.groups.filter(name='Gestor').exists(),
        'is_financeiro': user.groups.filter(name='Financeiro').exists(),
        'is_advogado': user.groups.filter(name='Advogado').exists(),
        'is_vendedor': user.groups.filter(name='Vendedor').exists(),
    }

def gestor_ou_admin_required(view_func):
    """ Decorator para as views de Gestão de Metas (movido para comissoes.views) """
    # (Este decorator foi movido para 'comissoes/views.py')
    # (Pode ser apagado daqui se não for mais usado em 'vendas')
    # *Vou deixá-lo aqui por segurança, caso o movamos de volta*
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        perms = _get_user_permissions(request.user)
        if not (perms['is_admin'] or perms['is_gestor']):
            messages.error(request, "Você não tem permissão para aceder a esta página.")
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return _wrapped_view

def _calcular_e_salvar_comissao(venda):
    """
    Executa a cascata de lógica de cálculo de comissão.
    (Agora importa 'RegraComissaoVendedor' de 'comissoes.models')
    """
    if venda.comissao_personalizada_valor is not None:
        venda.comissao_calculada_final = venda.comissao_personalizada_valor
        venda.save()
        return

    regra = RegraComissaoVendedor.objects.filter(vendedor=venda.vendedor, produto=venda.produto).first()
    if regra:
        if regra.tipo_comissao == 'F':
            venda.comissao_calculada_final = regra.valor_comissao
        else:
            venda.comissao_calculada_final = venda.honorarios * (regra.valor_comissao / Decimal(100))
        venda.save()
        return

    produto = venda.produto
    if produto.tipo_comissao == 'F':
        venda.comissao_calculada_final = produto.valor_comissao
    else:
        venda.comissao_calculada_final = venda.honorarios * (produto.valor_comissao / Decimal(100))
    venda.save()
    return

def _get_vendas_filtradas(request):
    """
    Função auxiliar que lê os filtros do request (GET) e retorna
    o queryset de Vendas já filtrado e com as permissões corretas.
    (Sem alterações de lógica, apenas de imports)
    """
    perms = _get_user_permissions(request.user)
    
    query_cliente = request.GET.get('cliente', '')
    query_vendedor = request.GET.get('vendedor', '')
    query_data_inicio = request.GET.get('data_inicio', '')
    query_data_fim = request.GET.get('data_fim', '')
    query_produto = request.GET.get('produto', '')
    query_status_venda = request.GET.get('status_venda', '')
    query_status_pagamento = request.GET.get('status_pagamento', '')
    query_status_contrato = request.GET.get('status_contrato', '')

    if perms['is_admin'] or perms['is_gestor'] or perms['is_financeiro'] or perms['is_advogado']:
        vendas = Venda.objects.all()
    else: 
        vendas = Venda.objects.filter(vendedor=request.user)

    if query_cliente:
        vendas = vendas.filter(cliente__nome_completo__icontains=query_cliente)
    if (perms['is_admin'] or perms['is_gestor'] or perms['is_financeiro'] or perms['is_advogado']) and query_vendedor:
        vendas = vendas.filter(vendedor_id=query_vendedor)
    if query_data_inicio:
        vendas = vendas.filter(data_venda__gte=query_data_inicio)
    if query_data_fim:
        try:
            data_fim_plus_one = datetime.strptime(query_data_fim, '%Y-%m-%d').date() + timedelta(days=1)
            vendas = vendas.filter(data_venda__lt=data_fim_plus_one)
        except (ValueError, TypeError): 
            pass
    if query_produto:
        vendas = vendas.filter(produto_id=query_produto)
    if query_status_venda:
        vendas = vendas.filter(status_venda=query_status_venda)
    if query_status_pagamento:
        vendas = vendas.filter(status_pagamento=query_status_pagamento)
    if query_status_contrato:
        vendas = vendas.filter(status_contrato=query_status_contrato)
        
    return vendas.select_related('cliente', 'produto', 'vendedor').order_by('-data_venda')

# ---
# SEÇÃO 2: VIEWS DE AUTENTICAÇÃO
# ---

def login_view(request):
    """ Página de login personalizada (/login/) """
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('dashboard')
    else:
        form = AuthenticationForm()
    return render(request, 'login.html', {'form': form, 'next': request.GET.get('next', '/')})

@login_required
def logout_view(request):
    """ Página de logout (/logout/) """
    logout(request)
    return redirect('login') 

# ---
# SEÇÃO 3: VIEWS DAS ABAS PRINCIPAIS (Vendas)
# ---

@login_required
def dashboard_graficos(request):
    """ Aba 1: Dashboard (Gráficos, KPIs e METAS) """
    perms = _get_user_permissions(request.user)
    
    # --- CORREÇÃO DE FUSO HORÁRIO (AQUI) ---
    # Usamos localdate() para obter a data na 'America/Sao_Paulo', 
    # em vez de .now().date() que usa UTC.
    today_date = timezone.localdate()
    # --- FIM DA CORREÇÃO ---
    
    default_start_str = (today_date - timedelta(days=30)).strftime('%Y-%m-%d')
    default_end_str = today_date.strftime('%Y-%m-%d')
    
    vendas_filtradas = _get_vendas_filtradas(request)
        
    kpis = vendas_filtradas.aggregate(
        total_vendas=Sum('honorarios'),
        ticket_medio=Avg('honorarios'),
        contagem_vendas=Count('id')
    )
    
    # Dados para Gráficos (Querysets)
    vendas_por_vendedor = vendas_filtradas.values('vendedor__username').annotate(total=Sum('honorarios')).order_by('-total')
    vendas_por_produto = vendas_filtradas.values('produto__nome').annotate(total=Sum('honorarios')).order_by('-total')
    vendas_por_mes = vendas_filtradas.annotate(mes=TruncMonth('data_venda')).values('mes').annotate(total=Sum('honorarios')).order_by('mes')
                           
    vendas_por_vendedor_payload = [{'vendedor__username': v['vendedor__username'], 'total': float(v['total'] or 0)} for v in vendas_por_vendedor]
    vendas_por_produto_payload = [{'produto__nome': p['produto__nome'], 'total': float(p['total'] or 0)} for p in vendas_por_produto]
    labels_linha = [v['mes'].strftime('%b/%Y') for v in vendas_por_mes]
    data_linha = [float(v['total'] or 0) for v in vendas_por_mes]
    
    # --- LÓGICA DE METAS ---
    grupos_do_utilizador = request.user.groups.all()
    
    if perms['is_admin'] or perms['is_gestor'] or perms['is_financeiro'] or perms['is_advogado']:
        # Admin/Gestor vê TODAS as metas ativas
        metas_ativas_hoje = MetaVenda.objects.filter(
            data_inicio__lte=today_date,
            data_fim__gte=today_date
        ).select_related('vendedor', 'grupo').order_by('grupo', 'vendedor')
    else:
        # Vendedor vê as suas + Gerais + Equipas
        filtro_metas_visiveis = Q(vendedor=request.user) | Q(vendedor__isnull=True, grupo__isnull=True) | Q(grupo__in=grupos_do_utilizador)
        metas_ativas_hoje = MetaVenda.objects.filter(
            filtro_metas_visiveis,
            data_inicio__lte=today_date,
            data_fim__gte=today_date
        ).select_related('vendedor', 'grupo').order_by('grupo', 'vendedor')

    # 2. Prepara os dados de progresso (O resto da lógica permanece 100% igual)
    meta_data_list = []
    
    for meta in metas_ativas_hoje:
        if meta.vendedor:
            filtro_vendas_meta = Q(vendedor=meta.vendedor)
            vendedor_nome = meta.vendedor.get_full_name() or meta.vendedor.username
            titulo = f"Meta Individual ({vendedor_nome})"
        elif meta.grupo:
            filtro_vendas_meta = Q(vendedor__groups=meta.grupo)
            titulo = f"Meta Equipa ({meta.grupo.name})"
        else:
            filtro_vendas_meta = Q() 
            titulo = "Meta Geral da Empresa"
        
        try:
            data_fim_para_query = meta.data_fim + timedelta(days=1)
        except TypeError:
            data_fim_para_query = meta.data_fim 

        progresso = Venda.objects.filter(
            filtro_vendas_meta,
            data_venda__gte=meta.data_inicio,
            data_venda__lt=data_fim_para_query, 
            status_pagamento='aprovado'
        ).aggregate(total=Sum('valor_entrada'))['total'] or 0

        percentual = int((progresso / meta.valor_meta) * 100) if meta.valor_meta > 0 else 0

        meta_data_list.append({
            'meta_titulo': titulo,
            'meta_periodo': f"{meta.data_inicio.strftime('%d/%m')} - {meta.data_fim.strftime('%d/%m')}",
            'meta_valor': meta.valor_meta,
            'meta_progresso': progresso,
            'meta_percentual': percentual
        })
    # --- FIM LÓGICA DE METAS ---
    
    todos_vendedores = User.objects.filter(is_superuser=False, is_active=True).order_by('username')
    todos_produtos = Produto.objects.all().order_by('nome') 

    context = {
        'perms': perms,
        'kpi_total_vendas': kpis.get('total_vendas') or 0,
        'kpi_ticket_medio': kpis.get('ticket_medio') or 0,
        'kpi_contagem_vendas': kpis.get('contagem_vendas') or 0,
        
        'vendas_por_vendedor_json': json.dumps(vendas_por_vendedor_payload),
        'vendas_por_produto_json': json.dumps(vendas_por_produto_payload),
        'vendas_por_mes_labels_json': json.dumps(labels_linha),
        'vendas_por_mes_data_json': json.dumps(data_linha),
        
        'todos_vendedores': todos_vendedores,
        'todos_produtos': todos_produtos,
        'status_venda_choices': Venda.STATUS_VENDA_CHOICES,
        'status_pagamento_choices': Venda.STATUS_PAGAMENTO_CHOICES,
        'status_contrato_choices': Venda.STATUS_CONTRATO_CHOICES,
        'query_data_inicio': request.GET.get('data_inicio', default_start_str), 
        'query_data_fim': request.GET.get('data_fim', default_end_str),
        'meta_data_list': meta_data_list,
    }
    return render(request, 'home.html', context)

@login_required
def lista_vendas(request):
    """ Aba 2: Lista de Vendas """
    perms = _get_user_permissions(request.user)
    vendas_filtradas = _get_vendas_filtradas(request)
    
    todos_vendedores = User.objects.filter(is_superuser=False, is_active=True).order_by('username')
    todos_produtos = Produto.objects.all().order_by('nome') # (Importa Produto de 'common.models')
    context = {
        'vendas': vendas_filtradas, 
        'perms': perms,
        'todos_vendedores': todos_vendedores, 
        'todos_produtos': todos_produtos,
        'status_venda_choices': Venda.STATUS_VENDA_CHOICES,
        'status_pagamento_choices': Venda.STATUS_PAGAMENTO_CHOICES,
        'status_contrato_choices': Venda.STATUS_CONTRATO_CHOICES,
    }
    return render(request, 'vendas/lista_vendas.html', context)

@login_required
@transaction.atomic 
def nova_venda(request):
    """ Aba 3: Formulário de Nova Venda (Usa ClienteForm de 'common.forms') """
    if request.method == 'POST':
        venda_form = VendaForm(request.POST)
        anexo_form = AnexoForm(request.POST, request.FILES)
        anexo_form.fields['arquivo'].required = False 
        
        cpf_cnpj = request.POST.get('cpf_cnpj', '').replace(r'\D', '')
        cliente_existente = Cliente.objects.filter(cpf_cnpj=cpf_cnpj).first()
        
        if cliente_existente:
            cliente_form = ClienteForm(request.POST, instance=cliente_existente)
        else:
            cliente_form = ClienteForm(request.POST)
        
        if cliente_form.is_valid() and venda_form.is_valid() and anexo_form.is_valid():
            cliente = cliente_form.save()
            venda = venda_form.save(commit=False)
            venda.cliente = cliente 
            venda.vendedor = request.user
            venda.save() 
            
            ficheiros_enviados = request.FILES.getlist('arquivo') 
            if ficheiros_enviados:
                for f in ficheiros_enviados:
                    AnexoVenda.objects.create(venda=venda, tipo='comprovante', arquivo=f, descricao=f.name)
                if venda.status_pagamento == 'pendente':
                    venda.status_pagamento = 'aguardando_validacao'
                    venda.save()
            return redirect('dashboard')
    
    else: # GET
        venda_form = VendaForm()
        cliente_form = ClienteForm() # (Importa ClienteForm de 'common.forms')
        anexo_form = AnexoForm()
        anexo_form.fields['arquivo'].required = False
        anexo_form.fields['arquivo'].label = "Anexar Comprovante(s) (Opcional)"
        
    context = {
        'venda_form': venda_form,
        'cliente_form': cliente_form,
        'anexo_form': anexo_form, 
    }
    return render(request, 'vendas/nova_venda.html', context)

@login_required
@transaction.atomic
def detalhe_venda(request, venda_id):
    """ Página de Detalhes da Venda (Usa ClienteEditForm de 'common.forms') """
    
    perms = _get_user_permissions(request.user)
    
    if perms['is_admin'] or perms['is_gestor'] or perms['is_financeiro'] or perms['is_advogado']:
        venda = get_object_or_404(Venda, id=venda_id)
    else: 
        venda = get_object_or_404(Venda, id=venda_id, vendedor=request.user)
    
    is_locked = (venda.status_contrato != 'nao_gerado')
    pagamento_aprovado = (venda.status_pagamento == 'aprovado')
    can_edit_data = (perms['is_admin'] or perms['is_gestor'] or (perms['is_vendedor'] and not is_locked))
    can_delete_comprovante = (perms['is_admin'] or perms['is_gestor'] or (perms['is_financeiro'] and not pagamento_aprovado) or (perms['is_vendedor'] and venda.vendedor == request.user and not pagamento_aprovado))
    can_delete_contrato = (perms['is_admin'] or perms['is_gestor'] or perms['is_advogado'])
    can_delete_nf = (perms['is_admin'] or perms['is_gestor'] or perms['is_financeiro'])

    if request.method == 'POST':
        acao = request.POST.get('acao')
        
        if acao in ['add_comprovante', 'add_contrato', 'add_nota_fiscal']:
            # ... (Lógica de upload de AnexoVenda permanece a mesma) ...
            tipo_anexo = acao.split('_')[-1]
            ficheiros_enviados = request.FILES.getlist('arquivos')
            pode_enviar = (perms['is_admin'] or perms['is_gestor'] or
                           (perms['is_financeiro'] and tipo_anexo in ['comprovante', 'nota_fiscal']) or
                           (perms['is_advogado'] and tipo_anexo == 'contrato') or
                           (perms['is_vendedor'] and tipo_anexo == 'comprovante'))
            
            if pode_enviar and ficheiros_enviados:
                for f in ficheiros_enviados:
                    AnexoVenda.objects.create(venda=venda, tipo=tipo_anexo, arquivo=f, descricao=f.name)
                if tipo_anexo == 'comprovante' and venda.status_pagamento == 'pendente':
                    venda.status_pagamento = 'aguardando_validacao'
                    venda.save()
                messages.success(request, f"{len(ficheiros_enviados)} ficheiro(s) enviado(s).")
            elif not ficheiros_enviados: messages.error(request, "Nenhum ficheiro selecionado.")
            else: messages.error(request, "Você não tem permissão para enviar este tipo de anexo.")
            return redirect('detalhe_venda', venda_id=venda.id)

        elif acao == 'update_status':
            status_pagamento_antigo = venda.status_pagamento 
            form = None
            if perms['is_admin'] or perms['is_gestor']: form = VendaStatusAdminForm(request.POST, instance=venda)
            elif perms['is_financeiro']: form = VendaStatusFinanceiroForm(request.POST, instance=venda)
            elif perms['is_advogado']: form = VendaStatusAdvogadoForm(request.POST, instance=venda)
            elif perms['is_vendedor'] and not is_locked: form = VendaStatusVendedorForm(request.POST, instance=venda)
            
            if form and form.is_valid():
                venda_atualizada = form.save()
                if venda_atualizada.status_pagamento == 'aprovado' and status_pagamento_antigo != 'aprovado':
                    _calcular_e_salvar_comissao(venda_atualizada) # (Função helper local)
                    messages.success(request, "Status atualizado e comissão calculada!")
                else:
                    messages.success(request, "Status da venda atualizado.")
            elif form: messages.error(request, "Erro ao atualizar o status.")
            else: messages.error(request, "Você não tem permissão para esta ação.")
            return redirect('detalhe_venda', venda_id=venda.id)

        elif acao == 'update_venda' or acao == 'update_cliente':
            if can_edit_data:
                if acao == 'update_venda': 
                    form = VendaEditForm(request.POST, instance=venda)
                    msg = "Dados da venda atualizados."
                else: 
                    form = ClienteEditForm(request.POST, instance=venda.cliente) # (Usa ClienteEditForm de 'common.forms')
                    msg = "Dados do cliente atualizados."
                
                if form.is_valid():
                    form.save()
                    if venda.status_pagamento == 'aprovado':
                        _calcular_e_salvar_comissao(venda)
                        msg += " Comissão recalculada."
                    messages.success(request, msg)
                else: messages.error(request, "Erro ao salvar. Verifique os campos.")
            else: messages.error(request, "Você não tem permissão para editar estes dados.")
            return redirect('detalhe_venda', venda_id=venda.id)

    if perms['is_admin'] or perms['is_gestor']: status_form = VendaStatusAdminForm(instance=venda)
    elif perms['is_financeiro']: status_form = VendaStatusFinanceiroForm(instance=venda)
    elif perms['is_advogado']: status_form = VendaStatusAdvogadoForm(instance=venda)
    else: status_form = VendaStatusVendedorForm(instance=venda)
    
    venda_form = VendaEditForm(instance=venda)
    cliente_form = ClienteEditForm(instance=venda.cliente) # (Usa ClienteEditForm de 'common.forms')

    context = {
        'venda': venda, 'anexo_form': AnexoForm(), 'status_form': status_form,
        'venda_form': venda_form, 'cliente_form': cliente_form,
        'perms': perms, 'is_locked': is_locked, 'pagamento_aprovado': pagamento_aprovado,
        'can_edit_data': can_edit_data,
        'can_delete_comprovante': can_delete_comprovante,
        'can_delete_contrato': can_delete_contrato,
        'can_delete_nf': can_delete_nf,
    }
    return render(request, 'vendas/detalhe_venda.html', context)

# ---
# SEÇÃO 4: VIEWS DE COMISSÕES (MOVIDAS PARA 'comissoes/views.py')
# ---
# (As views 'comissoes_dashboard', 'comissoes_historico_lotes', 
# 'comissoes_fechamento', 'comissoes_lote_detalhe' foram MOVIDAS)

# ---
# SEÇÃO 5: APIs INTERNAS (JSON) (MOVIDAS PARA 'common/views.py')
# ---
# (As views 'get_produto_data' e 'check_cliente' foram MOVIDAS)

# (A view 'delete_anexo' permanece aqui, pois está ligada ao AnexoVenda)
@login_required
@transaction.atomic
def delete_anexo(request, anexo_id):
    if request.method != 'POST': return HttpResponseForbidden()
    anexo = get_object_or_404(AnexoVenda, id=anexo_id); 
    venda = anexo.venda; 
    perms = _get_user_permissions(request.user); 
    pagamento_aprovado = (venda.status_pagamento == 'aprovado'); 
    pode_excluir = False
    if perms['is_admin'] or perms['is_gestor']:
        pode_excluir = True
    elif perms['is_financeiro']:
        if anexo.tipo == 'nota_fiscal': pode_excluir = True
        elif anexo.tipo == 'comprovante' and not pagamento_aprovado: pode_excluir = True
    elif perms['is_advogado'] and anexo.tipo == 'contrato':
        pode_excluir = True
    elif perms['is_vendedor'] and venda.vendedor == request.user:
        if anexo.tipo == 'comprovante' and not pagamento_aprovado:
            pode_excluir = True
    if pode_excluir:
        anexo.arquivo.delete(save=False); anexo.delete()
        messages.success(request, f"Anexo '{anexo.descricao or anexo.id}' foi excluído com sucesso.")
    else: messages.error(request, "Você não tem permissão para excluir este anexo.")
    return redirect('detalhe_venda', venda_id=venda.id)

# ---
# SEÇÃO 6: VIEWS DE EXPORTAÇÃO (RELATÓRIOS)
# ---
# (As views 'export_lotes_csv' e 'export_lotes_xlsx' foram MOVIDAS para 'comissoes/views.py')

@login_required
def export_vendas_csv(request):
    """ Exportação de Vendas (Permanece aqui) """
    vendas = _get_vendas_filtradas(request)
    response = HttpResponse(content_type='text/csv'); 
    response['Content-Disposition'] = 'attachment; filename="relatorio_vendas.csv"'
    writer = csv.writer(response, delimiter=';') 
    writer.writerow([
        'ID Venda', 
        'Data', 
        'Vendedor', 
        'Cliente', 
        'CPF/CNPJ', 
        'Produto', 
        'Honorarios Totais', 
        'Entrada', 
        'N Parcelas', 
        'Valor Parcela', 
        'Exito', 
        'Aporte', 
        'Status Venda', 
        'Status Pagamento', 
        'Status Contrato'
    ])
    for venda in vendas: 
        writer.writerow([
            venda.id, 
            venda.data_venda.strftime('%Y-%m-%d %H:%M'), 
            venda.vendedor.username, 
            venda.cliente.nome_completo, 
            venda.cliente.cpf_cnpj, 
            venda.produto.nome, 
            venda.honorarios, 
            venda.valor_entrada, 
            venda.num_parcelas, 
            venda.valor_parcela, 
            venda.valor_exito, 
            venda.valor_aporte, 
            venda.get_status_venda_display(), 
            venda.get_status_pagamento_display(), 
            venda.get_status_contrato_display()
        ])
    return response

@login_required
def export_vendas_xlsx(request):
    """ Exportação de Vendas (Permanece aqui) """
    vendas = _get_vendas_filtradas(request)
    wb = openpyxl.Workbook(); 
    ws = wb.active; ws.title = "Relatório de Vendas"
    headers = [
        'ID Venda',
        'Data',
        'Vendedor',
        'Cliente',
        'CPF/CNPJ',
        'Produto',
        'Honorarios Totais',
        'Entrada',
        'N Parcelas',
        'Valor Parcela',
        'Exito',
        'Aporte',
        'Status Venda',
        'Status Pagamento',
        'Status Contrato'
    ]
    ws.append(headers)
    for venda in vendas: 
        ws.append([
            venda.id,
            venda.data_venda.replace(tzinfo=None),
            venda.vendedor.username,
            venda.cliente.nome_completo,
            venda.cliente.cpf_cnpj,
            venda.produto.nome,
            float(venda.honorarios or 0),
            float(venda.valor_entrada or 0),
            venda.num_parcelas,
            float(venda.valor_parcela or 0),
            float(venda.valor_exito or 0),
            float(venda.valor_aporte or 0),
            venda.get_status_venda_display(),
            venda.get_status_pagamento_display(),
            venda.get_status_contrato_display()
        ])
    try:
        for col_num, header in enumerate(headers, 1): 
            ws.column_dimensions[get_column_letter(col_num)].width = 20
    except ImportError:
        pass 
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet');
    response['Content-Disposition'] = 'attachment; filename="relatorio_vendas.xlsx"'
    wb.save(response)
    return response

# ---
# SEÇÃO 7: VIEWS DE GESTÃO DE METAS (MOVIDAS PARA 'comissoes/views.py')
# ---
# (As views 'lista_metas', 'criar_meta', 'editar_meta', 'apagar_meta' foram MOVIDAS)